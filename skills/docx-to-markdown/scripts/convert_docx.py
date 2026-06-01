#!/usr/bin/env python3
"""
将docx文档转换为markdown格式，并提取所有图片到assets文件夹
支持将嵌入的Excel表格转换为Markdown表格
"""

import hashlib
import logging
import math
import os
import sys
import zipfile
import re
import xml.etree.ElementTree as ET
import io
import unicodedata
from html import unescape
from html.parser import HTMLParser
from collections import defaultdict
import posixpath
from typing import Dict, List, Optional

logger = logging.getLogger(__name__)


_FORBIDDEN_FILENAME_CHARS_RE = re.compile(r'[\\/:*?"<>|]')
_WHITESPACE_RE = re.compile(r"\s+")
_QUOTE_CHARS = '"“”‘’‚‛„‟«»‹›'


def _normalize_markdown_cell_text(value: str) -> str:
    """将单元格内容规范化为 Markdown 管道表可安全呈现的单行文本。"""
    text = unescape(value or "").replace("\xa0", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.split("\n")]
    lines = [line for line in lines if line]
    text = "<br>".join(lines) if lines else ""
    return text.replace("|", r"\|")


class _TableHTMLParser(HTMLParser):
    """解析单个 HTML table，保留 rowspan/colspan 与单元格文本。"""

    def __init__(self):
        super().__init__()
        self.rows = []
        self._in_tr = False
        self._in_cell = False
        self._current_row = []
        self._cell_parts = []
        self._cell_tag = None
        self._cell_rowspan = 1
        self._cell_colspan = 1

    @staticmethod
    def _safe_int(raw, default=1):
        try:
            value = int(raw)
            return value if value > 0 else default
        except Exception:
            return default

    def handle_starttag(self, tag, attrs):
        attrs_map = dict(attrs)
        tag = tag.lower()
        if tag == "tr":
            self._in_tr = True
            self._current_row = []
            return

        if tag in ("td", "th") and self._in_tr:
            self._in_cell = True
            self._cell_tag = tag
            self._cell_parts = []
            self._cell_rowspan = self._safe_int(attrs_map.get("rowspan"), 1)
            self._cell_colspan = self._safe_int(attrs_map.get("colspan"), 1)
            return

        if self._in_cell and tag in ("br",):
            self._cell_parts.append("\n")
        elif self._in_cell and tag in ("p", "div", "li"):
            if self._cell_parts and not self._cell_parts[-1].endswith("\n"):
                self._cell_parts.append("\n")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("p", "div", "li") and self._in_cell:
            if self._cell_parts and not self._cell_parts[-1].endswith("\n"):
                self._cell_parts.append("\n")
            return

        if tag in ("td", "th") and self._in_cell:
            text = _normalize_markdown_cell_text("".join(self._cell_parts))
            self._current_row.append(
                {
                    "text": text,
                    "rowspan": self._cell_rowspan,
                    "colspan": self._cell_colspan,
                    "is_header": self._cell_tag == "th",
                }
            )
            self._in_cell = False
            self._cell_parts = []
            self._cell_tag = None
            self._cell_rowspan = 1
            self._cell_colspan = 1
            return

        if tag == "tr" and self._in_tr:
            if self._current_row:
                self.rows.append(self._current_row)
            self._in_tr = False
            self._current_row = []

    def handle_data(self, data):
        if self._in_cell:
            self._cell_parts.append(data)


def _normalize_list_item_text(value: str) -> str:
    lines = []
    for raw in (value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n"):
        if not raw.strip():
            continue
        # 嵌套列表行保留原始缩进，避免层级被破坏。
        if re.match(r"^\s+(-|\d+\.)\s+", raw):
            lines.append(raw.rstrip())
        else:
            lines.append(re.sub(r"\s+", " ", raw).strip())
    return "\n".join(lines)


class _ListHTMLTransformer(HTMLParser):
    """将 HTML 列表结构转换为 Markdown 列表，保留嵌套层级。"""

    def __init__(self):
        super().__init__()
        self._out = []
        self._list_stack = []  # [{"type": "ul"/"ol", "items": [str, ...]}]
        self._li_stack = []  # [list[str], ...]

    @staticmethod
    def _attrs_to_str(attrs):
        if not attrs:
            return ""
        pairs = []
        for k, v in attrs:
            if v is None:
                pairs.append(k)
            else:
                escaped = str(v).replace('"', "&quot;")
                pairs.append(f'{k}="{escaped}"')
        return " " + " ".join(pairs)

    @staticmethod
    def _render_list(context, depth):
        indent = "  " * depth
        is_ordered = context["type"] == "ol"
        lines = []
        for idx, item in enumerate(context["items"], 1):
            marker = f"{idx}. " if is_ordered else "- "
            normalized = _normalize_list_item_text(item)
            if not normalized:
                lines.append(f"{indent}{marker}".rstrip())
                continue
            item_lines = normalized.splitlines()
            lines.append(f"{indent}{marker}{item_lines[0].strip()}")
            for extra in item_lines[1:]:
                if extra.startswith("  "):
                    lines.append(f"{indent}{extra}")
                else:
                    lines.append(f"{indent}  {extra.strip()}")
        return "\n".join(lines)

    def handle_starttag(self, tag, attrs):
        tag = tag.lower()
        if tag in ("ul", "ol"):
            self._list_stack.append({"type": tag, "items": []})
            return
        if tag == "li" and self._list_stack:
            self._li_stack.append([])
            return

        if self._li_stack:
            if tag == "br":
                self._li_stack[-1].append("\n")
            elif tag in ("p", "div"):
                if self._li_stack[-1] and not self._li_stack[-1][-1].endswith("\n"):
                    self._li_stack[-1].append("\n")
            return

        self._out.append(f"<{tag}{self._attrs_to_str(attrs)}>")

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in ("ul", "ol") and self._list_stack:
            context = self._list_stack.pop()
            md = self._render_list(context, len(self._list_stack))
            if self._li_stack:
                if self._li_stack[-1] and not self._li_stack[-1][-1].endswith("\n"):
                    self._li_stack[-1].append("\n")
                self._li_stack[-1].append(md)
            else:
                # 顶层列表后补空行，避免后续表格被当作列表延续文本。
                self._out.append("\n" + md + "\n\n")
            return

        if tag == "li" and self._li_stack:
            item_text = "".join(self._li_stack.pop())
            if self._list_stack:
                self._list_stack[-1]["items"].append(item_text)
            return

        if self._li_stack:
            if tag in ("p", "div"):
                if self._li_stack[-1] and not self._li_stack[-1][-1].endswith("\n"):
                    self._li_stack[-1].append("\n")
            return

        self._out.append(f"</{tag}>")

    def handle_startendtag(self, tag, attrs):
        tag = tag.lower()
        if self._li_stack and tag == "br":
            self._li_stack[-1].append("\n")
            return
        if self._li_stack:
            return
        self._out.append(f"<{tag}{self._attrs_to_str(attrs)}/>")

    def handle_data(self, data):
        if self._li_stack:
            self._li_stack[-1].append(data)
            return
        if self._list_stack:
            # 列表容器中但不在 li 内的噪声文本通常只有空白，忽略。
            return
        self._out.append(data)

    def get_output(self):
        return "".join(self._out)


def transform_html_lists_to_markdown(html: str) -> str:
    parser = _ListHTMLTransformer()
    parser.feed(html)
    parser.close()
    return parser.get_output()


def _expand_table_rows(rows):
    """将包含 rowspan/colspan 的行展开为等宽二维表。"""
    expanded = []
    spans = {}  # col_idx -> {"rows_left": int, "text": str}

    for row in rows:
        out_row = []
        col = 0

        def consume_span_at_current_col():
            nonlocal col
            while col in spans:
                span = spans[col]
                out_row.append(span["text"])
                span["rows_left"] -= 1
                if span["rows_left"] <= 0:
                    spans.pop(col, None)
                col += 1

        consume_span_at_current_col()
        for cell in row:
            consume_span_at_current_col()
            text = cell["text"]
            rowspan = max(1, int(cell["rowspan"]))
            colspan = max(1, int(cell["colspan"]))
            for offset in range(colspan):
                out_row.append(text)
                if rowspan > 1:
                    spans[col + offset] = {"rows_left": rowspan - 1, "text": text}
            col += colspan

        consume_span_at_current_col()
        expanded.append(out_row)

    while spans:
        out_row = []
        col = 0
        max_col = max(spans.keys())
        while col <= max_col:
            if col in spans:
                span = spans[col]
                out_row.append(span["text"])
                span["rows_left"] -= 1
                if span["rows_left"] <= 0:
                    spans.pop(col, None)
            else:
                out_row.append("")
            col += 1
        expanded.append(out_row)

    width = max((len(row) for row in expanded), default=0)
    if width:
        expanded = [row + [""] * (width - len(row)) for row in expanded]
    return expanded


def table_html_to_markdown(table_html: str) -> str:
    parser = _TableHTMLParser()
    parser.feed(table_html)
    parser.close()

    rows = _expand_table_rows(parser.rows)
    if not rows:
        return ""

    lines = []
    lines.append("| " + " | ".join(rows[0]) + " |")
    lines.append("| " + " | ".join(["---"] * len(rows[0])) + " |")
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines) + "\n\n"


def promote_numbered_bold_headings(markdown: str) -> str:
    """将“编号 + 加粗标题”段落提升为 Markdown 标题。"""
    pattern = re.compile(
        r"^(?P<num>\d+(?:\.\d+)*)(?P<dot>\.)?\s+\*\*(?P<title>[^*\n]+)\*\*\s*$",
        flags=0,
    )
    heading_pattern = re.compile(r"^(#{1,6})\s+")

    lines = markdown.splitlines()
    out = []
    previous_heading_level = 0
    previous_promoted_depth = None

    for line in lines:
        heading_match = heading_pattern.match(line)
        if heading_match:
            previous_heading_level = len(heading_match.group(1))
            previous_promoted_depth = None
            out.append(line)
            continue

        match = pattern.match(line.strip())
        if not match:
            out.append(line)
            continue

        num = match.group("num")
        dot = match.group("dot") or ""
        title = match.group("title").strip()
        depth = num.count(".") + 1
        level = min(depth, 6)  # 1级编号 -> #，2级编号 -> ##

        # 在深层章节下的“1. **小节**”更接近子标题，避免被抬到过高层级。
        if depth == 1 and previous_heading_level >= 2:
            if previous_promoted_depth == 1:
                level = previous_heading_level
            else:
                level = min(previous_heading_level + 1, 6)

        promoted = f"{'#' * level} {num}{dot} {title}"
        out.append(promoted)
        previous_heading_level = level
        previous_promoted_depth = depth

    # 保持原始编号，不做自动重排，避免双语并列标题或手工编号被误改。
    return "\n".join(out)


def promote_leading_bold_title(markdown: str) -> str:
    """将文档开头“整行加粗标题”提升为一级标题（保守触发）。"""
    lines = markdown.splitlines()
    first_idx = None
    for i, line in enumerate(lines):
        if line.strip():
            first_idx = i
            break
    if first_idx is None:
        return markdown

    first_line = lines[first_idx].strip()
    m = re.match(r"^\*\*(?P<title>.+?)\*\*$", first_line)
    if not m:
        return markdown

    # 仅在后续存在“编号章节标题”时触发，降低把普通强调段误判成标题的风险。
    section_heading_re = re.compile(r"^#{1,6}\s+\d+(?:\.\d+)*\.?\s+")
    has_numbered_section_heading = any(section_heading_re.match(line.strip()) for line in lines[first_idx + 1 :])
    if not has_numbered_section_heading:
        return markdown

    title = m.group("title").strip()
    if not title:
        return markdown

    lines[first_idx] = f"# {title}"
    return "\n".join(lines)


def sanitize_stem(stem: str) -> str:
    raw = stem  # 保留原始值用于 hash
    stem = unicodedata.normalize("NFKC", stem or "")
    for ch in _QUOTE_CHARS:
        stem = stem.replace(ch, "")
    stem = _FORBIDDEN_FILENAME_CHARS_RE.sub("_", stem)
    stem = _WHITESPACE_RE.sub(" ", stem).strip()
    stem = stem.strip(". ").strip()
    if not stem:
        return "document"
    if len(stem) <= 120:
        return stem
    # 截断时附加原始全名的短 hash，避免不同长文件名映射到同一输出目录
    suffix = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:8]
    return f"{stem[:111]}_{suffix}"


def extract_heading_level_map(docx_path: str) -> Dict[str, int]:
    """解析 DOCX 的 heading bookmark 段落样式，映射为 Markdown 标题层级。"""
    ns_w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    tag_p = f"{{{ns_w}}}p"
    tag_ppr = f"{{{ns_w}}}pPr"
    tag_pstyle = f"{{{ns_w}}}pStyle"
    tag_bm = f"{{{ns_w}}}bookmarkStart"
    attr_name = f"{{{ns_w}}}name"
    attr_val = f"{{{ns_w}}}val"
    tag_t = f"{{{ns_w}}}t"

    def style_to_level(style_val: str) -> Optional[int]:
        if not style_val:
            return None
        raw = str(style_val).strip()
        m = re.search(r"(\d+)$", raw)
        if not m:
            return None
        n = int(m.group(1))
        if n <= 0:
            return None
        # style=1/2/3 对应一/二/三级标题。
        return min(n, 6)

    def infer_level_from_text(text: str) -> Optional[int]:
        m = re.match(r"^\s*(\d+(?:\.\d+)*)\s*\.?\s+", text or "")
        if not m:
            return None
        depth = m.group(1).count(".") + 1
        return min(depth, 6)

    level_map: Dict[str, int] = {}
    try:
        with zipfile.ZipFile(docx_path, "r") as zip_ref:
            doc_xml = ET.fromstring(zip_ref.read("word/document.xml"))
        for p in doc_xml.findall(f".//{tag_p}"):
            bm = p.find(f".//{tag_bm}")
            if bm is None:
                continue
            name = bm.get(attr_name)
            if not name or not name.startswith("heading_"):
                continue

            ppr = p.find(tag_ppr)
            style_val = ""
            if ppr is not None:
                pstyle = ppr.find(tag_pstyle)
                if pstyle is not None:
                    style_val = pstyle.get(attr_val, "")

            level = style_to_level(style_val)
            if level is None:
                text = "".join((t.text or "") for t in p.findall(f".//{tag_t}"))
                level = infer_level_from_text(text)
            if level is not None:
                level_map[name] = level
    except Exception:
        return {}

    return level_map


def resolve_part_path(target: str) -> str:
    """将 relationship target 解析为 docx zip 内的规范路径（如 word/media/image1.png）"""
    target = (target or "").replace("\\", "/").strip()
    if not target:
        return ""
    if target.startswith("/"):
        target = target[1:]
    if target.startswith("word/"):
        return posixpath.normpath(target)
    return posixpath.normpath(posixpath.join("word", target))


def parse_relationships(docx_path):
    """解析docx中的关系文件，找出Excel嵌入和对应预览图的映射。

    策略：
      1. 优先从 document.xml 中解析 <w:object> 节点，提取 OLEObject rId
         和 imagedata rId 的真实配对关系（最可靠）。
      2. 对方法1未覆盖的项，使用 "rId相邻" 启发式补全（兼容）。
    """
    excel_to_preview = {}  # Excel路径 -> 预览图路径
    preview_to_excel = {}  # 预览图路径 -> Excel路径
    ordered_pairs = []  # [(Excel路径, 预览图路径)]，按文档出现顺序

    # --- 公共：解析 rels 文件，建立 rId -> target 映射 ---
    NS_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
    relationships = {}  # rId -> {'type': ..., 'target': ...}

    with zipfile.ZipFile(docx_path, 'r') as zip_ref:
        try:
            rels_content = zip_ref.read('word/_rels/document.xml.rels')
            rels_root = ET.fromstring(rels_content)
            for rel in rels_root.findall(f'.//{{{NS_REL}}}Relationship'):
                rid = rel.get('Id')
                rel_type = rel.get('Type', '').split('/')[-1]
                target = rel.get('Target', '')
                relationships[rid] = {'type': rel_type, 'target': target}
        except Exception as e:
            logger.warning("解析关系文件失败: %s", e)
            return excel_to_preview, preview_to_excel, ordered_pairs

        # --- 方法1：从 document.xml 解析 OLE 对象的真实引用 ---
        NS_W = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
        NS_V = "urn:schemas-microsoft-com:vml"
        NS_O = "urn:schemas-microsoft-com:office:office"
        NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

        try:
            doc_xml = zip_ref.read('word/document.xml')
            doc_root = ET.fromstring(doc_xml)

            # 查找所有 <w:object> 节点（可能嵌套在 mc:AlternateContent 等下面）
            for obj_node in doc_root.iter(f'{{{NS_W}}}object'):
                ole_rid = None
                img_rid = None

                # <o:OLEObject r:id="rIdX" />
                for ole in obj_node.iter(f'{{{NS_O}}}OLEObject'):
                    ole_rid = ole.get(f'{{{NS_R}}}id')

                # <v:imagedata r:id="rIdY" />
                for imgdata in obj_node.iter(f'{{{NS_V}}}imagedata'):
                    img_rid = imgdata.get(f'{{{NS_R}}}id')

                if ole_rid and img_rid and ole_rid in relationships and img_rid in relationships:
                    ole_target = resolve_part_path(relationships[ole_rid]['target'])
                    img_target = resolve_part_path(relationships[img_rid]['target'])
                    if ole_target.lower().endswith('.xlsx'):
                        excel_to_preview[ole_target] = img_target
                        preview_to_excel[img_target] = ole_target
                        ordered_pairs.append((ole_target, img_target))
        except Exception:
            pass  # document.xml 解析失败不影响后续

        # --- 方法2（补全）：rId 相邻启发式，补全方法1未覆盖的 Excel ---
        def rid_sort_key(rid: str) -> int:
            m = re.fullmatch(r"rId(\d+)", rid or "")
            return int(m.group(1)) if m else 10**9

        sorted_rids = sorted(relationships.keys(), key=rid_sort_key)

        for i, rid in enumerate(sorted_rids):
            rel = relationships[rid]
            if rel['type'] == 'package' and rel['target'].lower().endswith('.xlsx'):
                excel_file = resolve_part_path(rel['target'])
                if excel_file in excel_to_preview:
                    continue  # 已被方法1覆盖，跳过
                if i + 1 < len(sorted_rids):
                    next_rid = sorted_rids[i + 1]
                    next_rel = relationships[next_rid]
                    if next_rel['type'] == 'image':
                        preview_file = resolve_part_path(next_rel['target'])
                        excel_to_preview[excel_file] = preview_file
                        preview_to_excel[preview_file] = excel_file
                        ordered_pairs.append((excel_file, preview_file))

    return excel_to_preview, preview_to_excel, ordered_pairs


def _format_cell_value(cell) -> str:
    """将 openpyxl 单元格值转换为友好的字符串表示。"""
    if cell is None:
        return ''
    import datetime as _dt
    if isinstance(cell, _dt.datetime):
        if cell.hour == 0 and cell.minute == 0 and cell.second == 0 and cell.microsecond == 0:
            return cell.strftime("%Y-%m-%d")
        return cell.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(cell, _dt.date):
        return cell.strftime("%Y-%m-%d")
    if isinstance(cell, _dt.time):
        return cell.strftime("%H:%M:%S")
    if isinstance(cell, float) and math.isfinite(cell) and cell.is_integer():
        return str(int(cell))
    return str(cell)


def excel_to_markdown(xlsx_data):
    """将Excel数据转换为Markdown表格（仅依赖 openpyxl，无需 pandas）"""
    try:
        import openpyxl

        def normalize_rows(raw_rows: List[List[str]]) -> List[List[str]]:
            if not raw_rows:
                return []

            rows = [r for r in raw_rows if any(c.strip() for c in r)]
            if not rows:
                return []

            col_count = max(len(r) for r in rows)
            rows = [r + [''] * (col_count - len(r)) for r in rows]
            non_empty_cols = [j for j in range(col_count) if any(rows[i][j].strip() for i in range(len(rows)))]
            if not non_empty_cols:
                return []
            return [[r[j] for j in non_empty_cols] for r in rows]

        def apply_merged_cells(ws, raw_rows: List[List[str]]) -> List[List[str]]:
            """将合并单元格展开为 Markdown 管道表可读的全展开网格。"""
            if not raw_rows:
                return raw_rows

            for merged in ws.merged_cells.ranges:
                min_row, max_row = merged.min_row, merged.max_row
                min_col, max_col = merged.min_col, merged.max_col

                row_idx = min_row - 1
                col_idx = min_col - 1
                if row_idx >= len(raw_rows):
                    continue
                if col_idx >= len(raw_rows[row_idx]):
                    continue

                anchor_value = raw_rows[row_idx][col_idx]
                if not anchor_value:
                    continue

                for row_no in range(min_row, max_row + 1):
                    i = row_no - 1
                    if i >= len(raw_rows):
                        continue
                    if len(raw_rows[i]) < max_col:
                        raw_rows[i].extend([''] * (max_col - len(raw_rows[i])))
                    for col_no in range(min_col, max_col + 1):
                        raw_rows[i][col_no - 1] = anchor_value
            return raw_rows

        def sheet_to_rows(ws) -> tuple[List[List[str]], List[str], int]:
            raw_rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                raw_rows.append([
                    _normalize_markdown_cell_text(_format_cell_value(cell))
                    for cell in row
                ])
            score_rows = normalize_rows(raw_rows)
            score = sum(1 for r in score_rows for c in r if c.strip())
            merge_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            raw_rows = apply_merged_cells(ws, raw_rows)
            return normalize_rows(raw_rows), merge_ranges, score

        # 不使用 read_only=True：部分嵌入工作簿的维度元数据异常，read_only 模式会把表格截断成 1x1。
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=False, data_only=True)
        best_rows = []
        best_merge_ranges: List[str] = []
        best_score = -1
        for ws in wb.worksheets:
            rows, merge_ranges, score = sheet_to_rows(ws)
            if not rows:
                continue
            if score > best_score:
                best_rows = rows
                best_merge_ranges = merge_ranges
                best_score = score
        wb.close()

        if not best_rows:
            return None

        header = '| ' + ' | '.join(best_rows[0]) + ' |'
        separator = '| ' + ' | '.join(['---'] * len(best_rows[0])) + ' |'
        body_lines = ['| ' + ' | '.join(r) + ' |' for r in best_rows[1:]]
        table_text = header + '\n' + separator + '\n' + '\n'.join(body_lines)
        if best_merge_ranges:
            ranges_text = ", ".join(best_merge_ranges)
            return f"> merge_ranges: {ranges_text}\n\n{table_text}"
        return table_text

    except Exception as e:
        logger.warning("Excel转Markdown失败: %s", e)
        return None


def detect_image_format(image_data):
    """检测图片的真实格式"""
    if image_data[:8] == b'\x89PNG\r\n\x1a\n':
        return '.png'
    elif image_data[:2] == b'\xff\xd8':
        return '.jpeg'
    elif image_data[:6] in (b'GIF87a', b'GIF89a'):
        return '.gif'
    elif image_data[:4] == b'RIFF' and image_data[8:12] == b'WEBP':
        return '.webp'
    elif image_data[:2] == b'BM':
        return '.bmp'
    else:
        return '.png'  # 默认


def extract_content_from_docx(docx_path, assets_dir):
    """从docx中提取图片和Excel数据，并构建“内容hash -> 内容”的映射

    返回:
        image_by_hash: { sha256_hex: "assets/xxx.png" }
        table_queue_by_hash: { sha256_hex: ["<md_table1>", "<md_table2>", ...] }
        table_repeat_by_hash: { sha256_hex: "<md_table>" }  # 队列耗尽时的稳定兜底
    """
    image_by_hash = {}
    table_queue_by_hash = defaultdict(list)
    table_repeat_by_hash = {}
    
    # 解析关系，找出Excel和预览图的对应
    excel_to_preview, preview_to_excel, ordered_pairs = parse_relationships(docx_path)
    
    with zipfile.ZipFile(docx_path, 'r') as zip_ref:
        excel_md_by_path = {}
        table_preview_paths = set()

        # 先提取所有 Excel 文件的数据并转换为 Markdown
        for file_info in zip_ref.filelist:
            if file_info.filename.startswith('word/embeddings/') and file_info.filename.lower().endswith('.xlsx'):
                excel_file = file_info.filename
                xlsx_data = zip_ref.read(file_info.filename)

                markdown_table = excel_to_markdown(xlsx_data)
                if markdown_table:
                    excel_md_by_path[excel_file] = markdown_table
                else:
                    logger.warning("Excel表格转换失败（将保留预览图）: %s", excel_file)

        # 建立预览图 hash -> 表格队列（同一预览图内容可对应多个表格）
        pairs = ordered_pairs if ordered_pairs else [(e, p) for e, p in excel_to_preview.items()]
        for excel_path, preview_path in pairs:
            table_md = excel_md_by_path.get(excel_path)
            if not table_md:
                continue
            if preview_path not in zip_ref.namelist():
                continue
            preview_data = zip_ref.read(preview_path)
            digest = hashlib.sha256(preview_data).hexdigest()
            table_queue_by_hash[digest].append(table_md)
            table_repeat_by_hash[digest] = table_md
            table_preview_paths.add(preview_path)
            logger.info("转换Excel为表格: %s", excel_path)

        # 处理图片
        for file_info in zip_ref.filelist:
            if file_info.filename.startswith('word/media/'):
                image_name = os.path.basename(file_info.filename)
                
                # 检查这个图片是否是Excel的预览图
                if file_info.filename in table_preview_paths:
                    continue
                
                # 普通图片，直接提取
                image_data = zip_ref.read(file_info.filename)
                digest = hashlib.sha256(image_data).hexdigest()
                
                # 检测真实的图片格式并修正扩展名
                actual_ext = detect_image_format(image_data)
                base_name = os.path.splitext(image_name)[0]
                corrected_name = f"{base_name}{actual_ext}"
                
                image_path = os.path.join(assets_dir, corrected_name)
                # 扩展名修正后可能与已有文件同名，若内容不同则附加hash后缀避免覆盖
                if os.path.exists(image_path):
                    try:
                        with open(image_path, "rb") as f:
                            existing = f.read()
                        if existing != image_data:
                            corrected_name = f"{base_name}_{digest[:8]}{actual_ext}"
                            image_path = os.path.join(assets_dir, corrected_name)
                    except Exception:
                        corrected_name = f"{base_name}_{digest[:8]}{actual_ext}"
                        image_path = os.path.join(assets_dir, corrected_name)

                if not os.path.exists(image_path):
                    with open(image_path, 'wb') as f:
                        f.write(image_data)

                image_by_hash.setdefault(digest, f"assets/{corrected_name}")
                logger.info("提取图片: %s", corrected_name)
    
    return image_by_hash, table_queue_by_hash, table_repeat_by_hash


def extract_textbox_content(docx_path: str) -> List[str]:
    """从 DOCX 的 document.xml 中提取文本框 (<w:txbxContent>) 内的纯文本。

    mammoth 通常会忽略 text box / shape 中的内容，此函数作为补充。
    返回非空文本块列表。
    """
    ns_w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    ns_wps = "http://schemas.microsoft.com/office/word/2010/wordprocessingShape"
    tag_txbx = f"{{{ns_w}}}txbxContent"
    tag_txbx_wps = f"{{{ns_wps}}}txbxContent"
    tag_t = f"{{{ns_w}}}t"
    tag_p = f"{{{ns_w}}}p"

    blocks: List[str] = []
    try:
        with zipfile.ZipFile(docx_path, "r") as zf:
            doc_xml = ET.fromstring(zf.read("word/document.xml"))

        for txbx_tag in (tag_txbx, tag_txbx_wps):
            for txbx in doc_xml.iter(txbx_tag):
                paras = []
                for p in txbx.findall(f".//{tag_p}"):
                    text = "".join((t.text or "") for t in p.findall(f".//{tag_t}"))
                    text = text.strip()
                    if text:
                        paras.append(text)
                if paras:
                    blocks.append("\n".join(paras))
    except Exception:
        pass
    return blocks


def extract_math_text(docx_path: str) -> List[str]:
    """从 DOCX 的 document.xml 中提取 OMML 数学公式的纯文本内容。

    完整的 OMML→LaTeX 转换极为复杂，此函数仅提取公式中的文本节点，
    用 $ 包裹作为占位标记，便于下游人工校正。
    """
    ns_m = "http://schemas.openxmlformats.org/officeDocument/2006/math"
    ns_w = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    tag_omath = f"{{{ns_m}}}oMath"
    tag_omath_para = f"{{{ns_m}}}oMathPara"
    tag_t_m = f"{{{ns_m}}}t"
    tag_t_w = f"{{{ns_w}}}t"

    formulas: List[str] = []
    try:
        with zipfile.ZipFile(docx_path, "r") as zf:
            doc_xml = ET.fromstring(zf.read("word/document.xml"))

        seen = set()
        for parent_tag in (tag_omath_para, tag_omath):
            for node in doc_xml.iter(parent_tag):
                node_id = id(node)
                if node_id in seen:
                    continue
                seen.add(node_id)
                parts = []
                for t in node.iter():
                    if t.tag in (tag_t_m, tag_t_w) and t.text:
                        parts.append(t.text)
                text = "".join(parts).strip()
                if text:
                    formulas.append(text)
                for child in node.iter(tag_omath):
                    seen.add(id(child))
    except Exception:
        pass
    return formulas


def convert_docx_to_markdown(docx_path, output_dir, create_subfolder=True, output_name=None):
    """将docx转换为markdown

    Args:
        docx_path: DOCX 文件路径
        output_dir: 输出目录路径
        create_subfolder: 是否在输出目录下创建以文件名命名的子文件夹（默认 True）
        output_name: 自定义输出文件名（不含扩展名），如不提供则从 docx_path 提取
    """

    # 先校验输入，避免 BadZipFile 直接中断并泄漏底层异常。
    try:
        with zipfile.ZipFile(docx_path, "r") as zip_ref:
            if "word/document.xml" not in zip_ref.namelist():
                raise ValueError(f"输入文件不是有效的 DOCX（缺少 word/document.xml）: {docx_path}")
    except zipfile.BadZipFile as exc:
        raise ValueError(f"输入文件不是有效的 DOCX/ZIP: {docx_path}") from exc

    # 获取文件名（不含扩展名）
    if output_name:
        base_name = os.path.splitext(output_name)[0]
    else:
        base_name = os.path.splitext(os.path.basename(docx_path))[0]
    folder_name = sanitize_stem(base_name)
    
    # 确定最终输出目录
    if create_subfolder:
        final_output_dir = os.path.join(output_dir, folder_name)
    else:
        final_output_dir = output_dir
    
    # 创建输出目录
    os.makedirs(final_output_dir, exist_ok=True)
    assets_dir = os.path.join(final_output_dir, 'assets')
    os.makedirs(assets_dir, exist_ok=True)
    
    # 提取图片和Excel表格
    logger.info("正在提取内容...")
    image_by_hash, table_queue_by_hash, table_repeat_by_hash = extract_content_from_docx(docx_path, assets_dir)
    table_md_by_placeholder = {}
    table_seq = [0]
    heading_level_map = extract_heading_level_map(docx_path)
    
    # 使用mammoth转换为HTML
    logger.info("正在转换文档...")

    def convert_image(image):
        """根据图片内容hash，返回对应的assets路径或表格占位符"""
        with image.open() as image_bytes:
            image_data = image_bytes.read()
        digest = hashlib.sha256(image_data).hexdigest()

        table_queue = table_queue_by_hash.get(digest)
        if table_queue:
            # 若仅剩一个元素则不再弹出，确保同一预览图多次出现时仍稳定替换为表格
            table_md = table_queue[0] if len(table_queue) == 1 else table_queue.pop(0)
            placeholder = f"__TABLE_PLACEHOLDER_{digest}_{table_seq[0]}__"
            table_seq[0] += 1
            table_md_by_placeholder[placeholder] = table_md
            return {"src": placeholder}

        # 防御性兜底：当前队列逻辑保证最后一个元素不会被弹出，因此此分支在
        # 正常流程中不会触发。保留此分支作为安全网，以防未来队列策略调整后
        # 队列被完全消耗的情况，确保仍能稳定替换为表格而非退化为普通图片。
        if digest in table_repeat_by_hash:
            table_md = table_repeat_by_hash[digest]
            placeholder = f"__TABLE_PLACEHOLDER_{digest}_{table_seq[0]}__"
            table_seq[0] += 1
            table_md_by_placeholder[placeholder] = table_md
            return {"src": placeholder}

        image_src = image_by_hash.get(digest)
        if image_src:
            return {"src": image_src}

        # 兜底：某些情况下zip里的图片与mammoth回调数据不一致，直接按hash写入assets
        ext = detect_image_format(image_data)
        filename = f"image_{digest[:16]}{ext}"
        image_path = os.path.join(assets_dir, filename)
        if not os.path.exists(image_path):
            with open(image_path, "wb") as f:
                f.write(image_data)
        image_by_hash[digest] = f"assets/{filename}"
        return {"src": f"assets/{filename}"}
    
    with open(docx_path, 'rb') as docx_file:
        import mammoth

        result = mammoth.convert_to_html(
            docx_file,
            convert_image=mammoth.images.img_element(convert_image)
        )
        html = result.value
        for msg in getattr(result, "messages", []) or []:
            logger.debug("mammoth提示: %s", msg)
    
    # 将HTML转换为Markdown
    markdown = html_to_markdown(html, heading_level_map)
    
    # 替换表格占位符
    for placeholder_key, table_md in table_md_by_placeholder.items():
        placeholder = f"![]({placeholder_key})"
        markdown = markdown.replace(placeholder, f"\n\n{table_md}\n\n")

    # 移除嵌入 Excel 替换后残留的预览图说明文本
    markdown = re.sub(
        r"\n+\*{0,2}点击图片可查看完整电子表格\*{0,2}\s*\n",
        "\n",
        markdown,
    )

    # 追加 mammoth 未能提取的文本框内容
    textbox_blocks = extract_textbox_content(docx_path)
    if textbox_blocks:
        # 检查主体中是否已包含文本框文本（mammoth 有时也能提取部分文本框）
        missing = [b for b in textbox_blocks if b.splitlines()[0] not in markdown]
        if missing:
            markdown += "\n\n---\n\n> **\\[文本框内容\\]**\n\n"
            for block in missing:
                markdown += f"> {block}\n>\n"
            logger.info("追加了 %d 个文本框内容", len(missing))

    # 追加 mammoth 未能提取的数学公式
    math_formulas = extract_math_text(docx_path)
    if math_formulas:
        missing_math = [f for f in math_formulas if f not in markdown]
        if missing_math:
            markdown += "\n\n---\n\n> **\\[数学公式\\]**\n\n"
            for formula in missing_math:
                markdown += f"> $$ {formula} $$\n>\n"
            logger.info("追加了 %d 个数学公式", len(missing_math))

    md_path = os.path.join(final_output_dir, f"{folder_name}.md")
    
    with open(md_path, 'w', encoding='utf-8') as f:
        f.write(markdown)
    
    logger.info("转换完成: %s", md_path)
    return md_path


def _convert_footnotes(html: str) -> str:
    """将 mammoth 生成的脚注 HTML 转换为 Markdown 脚注语法。

    mammoth 输出格式：
      正文引用: <sup><a href="#footnote-N" id="footnote-ref-N">[N]</a></sup>
      文末列表: <li id="footnote-N"><p>text <a href="#footnote-ref-N">↑</a></p></li>
    """
    footnote_bodies: Dict[str, str] = {}

    def _extract_footnote_body(match):
        fid = match.group("fid")
        body_html = match.group("body")
        body_html = re.sub(r"</(?:p|div|li|br)\s*/?>", " ", body_html, flags=re.IGNORECASE)
        body = re.sub(r"<[^>]+>", "", body_html, flags=re.DOTALL)
        body = unescape(body).replace("↑", "").strip()
        body = re.sub(r"  +", " ", body)
        if body:
            footnote_bodies[fid] = body
        return ""

    html = re.sub(
        r'<li\b[^>]*\bid\s*=\s*["\']?footnote-(?P<fid>\d+)["\']?[^>]*>'
        r"(?P<body>.*?)</li>",
        _extract_footnote_body,
        html,
        flags=re.DOTALL | re.IGNORECASE,
    )

    html = re.sub(
        r"<sup>\s*<a\b[^>]*href\s*=\s*[\"']?#footnote-(\d+)[\"']?[^>]*>"
        r"\s*\[\d+\]\s*</a>\s*</sup>",
        lambda m: f"[^{m.group(1)}]",
        html,
        flags=re.IGNORECASE,
    )

    if footnote_bodies:
        footer = "\n\n---\n\n"
        for fid in sorted(footnote_bodies, key=int):
            footer += f"[^{fid}]: {footnote_bodies[fid]}\n"
        html += footer

    return html


def html_to_markdown(html, heading_level_map: Optional[Dict[str, int]] = None):
    """将HTML转换为Markdown"""

    html = _convert_footnotes(html)

    # 处理标题
    html = re.sub(r'<h1[^>]*>(.*?)</h1>', r'# \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h2[^>]*>(.*?)</h2>', r'## \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h3[^>]*>(.*?)</h3>', r'### \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h4[^>]*>(.*?)</h4>', r'#### \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h5[^>]*>(.*?)</h5>', r'##### \1\n\n', html, flags=re.DOTALL)
    html = re.sub(r'<h6[^>]*>(.*?)</h6>', r'###### \1\n\n', html, flags=re.DOTALL)
    
    # 优先按 DOCX 原始 heading 样式提升标题层级（主流程），避免纯文本启发式误判。
    if heading_level_map:
        def _replace_anchored_heading(match):
            heading_id = match.group("id1") or match.group("id2") or match.group("id3") or ""
            content = match.group("content")
            level = heading_level_map.get(heading_id)
            if not level:
                return match.group(0)
            text = re.sub(r"<[^>]+>", "", content, flags=re.DOTALL)
            text = unescape(text).strip()
            if not text:
                return ""
            return f"{'#' * level} {text}\n\n"

        html = re.sub(
            (
                r"<p[^>]*>\s*"
                r"<a[^>]*\bid\s*=\s*(?:\"(?P<id1>heading_\d+)\"|'(?P<id2>heading_\d+)'|(?P<id3>heading_\d+))[^>]*>"
                r"\s*</a>(?P<content>.*?)</p>"
            ),
            _replace_anchored_heading,
            html,
            flags=re.DOTALL | re.IGNORECASE,
        )

    # 处理粗体和斜体
    html = re.sub(r'<strong>(.*?)</strong>', r'**\1**', html, flags=re.DOTALL)
    html = re.sub(r'<b>(.*?)</b>', r'**\1**', html, flags=re.DOTALL)
    html = re.sub(r'<em>(.*?)</em>', r'*\1*', html, flags=re.DOTALL)
    html = re.sub(r'<i>(.*?)</i>', r'*\1*', html, flags=re.DOTALL)
    
    # 处理图片
    def _replace_img(match):
        src = match.group("src1") or match.group("src2") or match.group("src3") or ""
        return f"![]({src})\n\n"

    html = re.sub(
        (
            r"<img\b[^>]*\bsrc\s*=\s*"
            r"(?:\"(?P<src1>[^\"]*)\"|'(?P<src2>[^']*)'|(?P<src3>[^\s\"'=<>`]+))"
            r"[^>]*/?>"
        ),
        _replace_img,
        html,
        flags=re.IGNORECASE,
    )
    
    # 处理链接（支持双引号、单引号、无引号三种 href 写法）
    def _replace_link(match):
        href = match.group("href1") or match.group("href2") or match.group("href3") or ""
        text = match.group("text")
        return f"[{text}]({href})"

    html = re.sub(
        (
            r"<a\b[^>]*\bhref\s*=\s*"
            r"(?:\"(?P<href1>[^\"]*)\"|'(?P<href2>[^']*)'|(?P<href3>[^\s\"'=<>`]+))"
            r"[^>]*>(?P<text>.*?)</a>"
        ),
        _replace_link,
        html,
        flags=re.DOTALL | re.IGNORECASE,
    )

    # 先把HTML里的换行标签转为文本换行（需早于表格转换，避免改写表格里的 <br> 文本）
    html = re.sub(r'<br\s*/?>', '\n', html)

    # 先处理表格（必须在段落/列表转换之前）
    html = re.sub(
        r'<table[^>]*>.*?</table>',
        lambda match: table_html_to_markdown(match.group(0)),
        html,
        flags=re.DOTALL | re.IGNORECASE,
    )

    # 使用结构化解析处理列表，避免正则顺序导致的嵌套层级破坏。
    html = transform_html_lists_to_markdown(html)
    
    # 处理段落
    html = re.sub(r'<p[^>]*>(.*?)</p>', r'\1\n\n', html, flags=re.DOTALL)
    
    # 移除剩余的HTML标签（保留 <br> 供 Markdown 单元格换行显示）
    html = re.sub(r'<(?!br\s*/?)[^>]+>', '', html, flags=re.IGNORECASE)
    
    # 清理多余的空行
    html = re.sub(r'\n{3,}', '\n\n', html)
    
    html = html.replace('&nbsp;', ' ')
    html = unescape(html)

    html = promote_numbered_bold_headings(html)
    html = promote_leading_bold_title(html)
    return html.strip()


if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format="%(message)s")

    if len(sys.argv) < 3:
        print("用法: python scripts/convert_docx.py <docx文件路径> <输出目录>  (在skill目录执行)")
        print("或:   python convert_docx.py <docx文件路径> <输出目录>          (在scripts目录执行)")
        sys.exit(1)
    
    docx_path = sys.argv[1]
    output_dir = sys.argv[2]
    
    if not os.path.exists(docx_path):
        logger.error("文件不存在 - %s", docx_path)
        sys.exit(1)
    
    convert_docx_to_markdown(docx_path, output_dir)
